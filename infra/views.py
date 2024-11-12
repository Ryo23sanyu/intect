from collections import defaultdict
from copy import copy
import datetime
import fnmatch
from io import BytesIO
import io
from itertools import groupby
import math
from operator import attrgetter
import os
import re
import tempfile
import time
from urllib.parse import unquote
import boto3
from django.conf import settings
from django.db import IntegrityError
from django.db.models import Count
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, DeleteView, UpdateView
from django.views import View
import urllib

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import requests

from infra.number_replace import process_inspection_data
from infra.picture_download import download_and_zip_images
from infra.tasks import create_picturelist

from .forms import DamageCommentCauseEditForm, DamageCommentEditForm, DamageCommentJadgementEditForm, NameEntryForm, PartsNumberForm, TableForm
from .models import Approach, Article, BridgePicture, DamageComment, DamageList, FullReportData, Infra, LoadGrade, LoadWeight, NameEntry, PartsName, PartsNumber, Regulation, Rulebook, Table, Thirdparty, UnderCondition

# << indexページ(使用方法) >>
def index_view(request):
    return render(request, 'infra/how_to_use.html')

# << 案件・一覧 >>
class ListArticleView(LoginRequiredMixin, ListView):
    template_name = 'infra/article_list.html'
    model = Article
    
# << 案件・詳細 >>
class DetailArticleView(LoginRequiredMixin, DetailView):
    template_name = 'infra/article_detail.html'
    model = Article
    
# << 案件・作成 >>
class CreateArticleView(LoginRequiredMixin, CreateView):
    template_name = 'infra/article_create.html'
    model = Article
    fields = ('案件名', '土木事務所', '対象数', '担当者名', 'その他')
    success_url = reverse_lazy('list-article')
    
# << 案件・削除 >>
class DeleteArticleView(LoginRequiredMixin, DeleteView):
    template_name = 'infra/article_delete.html'
    model = Article
    success_url = reverse_lazy('list-article')
    
# << 案件・更新 >>
class UpdateArticleView(LoginRequiredMixin, UpdateView):
    template_name = 'infra/article_update.html'
    model = Article
    fields = ('案件名', '土木事務所', '対象数', '担当者名', 'その他')
    success_url = reverse_lazy('list-article')


# << 橋梁・一覧 >>
class ListInfraView(LoginRequiredMixin, ListView):
    template_name = 'infra/infra_list.html'
    model = Infra # 使用するモデル「infra」
    def get_queryset(self, **kwargs):
        # モデル検索のクエリー。Infra.objects.all() と同じ結果で全ての Infra
        queryset = super().get_queryset(**kwargs)
        queryset = queryset.filter(article = self.kwargs["article_pk"])
        return queryset
    def get_context_data(self, **kwargs):
        kwargs["article_id"] = self.kwargs["article_pk"]
        return super().get_context_data(**kwargs)
      
      
# << 橋梁・詳細 >>
class DetailInfraView(LoginRequiredMixin, DetailView):
    template_name = 'infra/infra_detail.html'
    model = Infra
    def get_context_data(self, **kwargs):
        # HTMLテンプレートでの表示変数として「article_id」を追加。
        # 値はパスパラメータpkの値→取り扱うarticle.idとなる
        kwargs["article_id"] = self.kwargs["article_pk"]
        #モデルのTableクラス ↑                    ↑  infraに格納する値は自らのpkの値とする
        return super().get_context_data(**kwargs)
      
# << 橋梁・作成 >>
class CreateInfraView(LoginRequiredMixin, CreateView):
    template_name = 'infra/infra_create.html'
    model = Infra
    fields = ('title', '径間数', '橋長', '全幅員','橋梁コード', '活荷重', '等級', '適用示方書', '路線名',
              '上部構造形式', '下部構造形式', '基礎構造形式', '近接方法', '交通規制', '第三者点検', '海岸線との距離', 
              '路下条件', '交通量', '大型車混入率', '特記事項', 'カテゴリー', 'latitude', 'longitude', 'end_latitude', 'end_longitude')
    success_url = reverse_lazy('detail-infra')
    
    def form_valid(self, form): # form_validはフォームが有効のとき呼び出される
        article_pk = self.kwargs['article_pk'] # URLパラメータからarticle_pkを取得
        print(article_pk)
        article = get_object_or_404(Article, pk=article_pk) # article_pkを使って、Articleモデルから対応するオブジェクトを取得。オブジェクトが見つからない場合は404エラーを返す
        print(article)
        form.instance.article = article # フォームインスタンス (form.instance) の articleフィールドに取得したarticleをセット
        print(form.instance)
        self.article = article # インスタンス変数として保存
        return super().form_valid(form)
    def get_success_url(self):
        return reverse('list-infra', kwargs={'article_pk': self.article.pk})
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["loadWeights"] = LoadWeight.objects.all()
        context["loadGrades"] = LoadGrade.objects.all()
        context["rulebooks"] = Rulebook.objects.all()
        context["approachs"] = Approach.objects.all()
        context["regulations"] = Regulation.objects.all()
        context["thirdpartys"] = Thirdparty.objects.all()
        context["underconditions"] = UnderCondition.objects.all()
        return context
      
# << 橋梁・削除 >>
class DeleteInfraView(LoginRequiredMixin, DeleteView):
    template_name = 'infra/infra_delete.html'
    model = Infra
    success_url = reverse_lazy('list-infra')
    def get_success_url(self):
        return reverse_lazy('list-infra', kwargs={'article_pk': self.kwargs["article_pk"]})
      
# << 橋梁・更新 >>
class UpdateInfraView(LoginRequiredMixin, UpdateView):
    template_name = 'infra/infra_update.html'
    model = Infra
    fields = ('title', '径間数', '橋長', '全幅員', 'latitude', 'longitude', '橋梁コード', '活荷重', '等級', '適用示方書', 
              '上部構造形式', '下部構造形式', '基礎構造形式', '近接方法', '交通規制', '第三者点検', '海岸線との距離', 
              '路下条件', '交通量', '大型車混入率', '特記事項', 'カテゴリー', 'article')
    success_url = reverse_lazy('detail-infra')
    def get_success_url(self):
        return reverse_lazy('detail-infra', kwargs={'article_pk': self.kwargs["article_pk"], 'pk': self.kwargs["pk"]})

    #新規作成時、交通規制の全データをコンテキストに含める。
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_regulations = self.object.交通規制.values_list('id', flat=True)# 選択状態を保持
        context['selected_regulations'] = list(selected_regulations)# 選択状態を保持
        context["regulations"] = Regulation.objects.all()
        
        selected_loadWeights = self.object.活荷重.values_list('id', flat=True)
        context['selected_loadWeights'] = list(selected_loadWeights)
        context["loadWeights"] = LoadWeight.objects.all()
        
        selected_loadGrades = self.object.等級.values_list('id', flat=True)
        context['selected_loadGrades'] = list(selected_loadGrades)
        context["loadGrades"] = LoadGrade.objects.all()
        
        selected_rulebooks = self.object.適用示方書.values_list('id', flat=True)
        context['selected_rulebooks'] = list(selected_rulebooks)
        context["rulebooks"] = Rulebook.objects.all()
        
        selected_approachs = self.object.近接方法.values_list('id', flat=True)
        context['selected_approachs'] = list(selected_approachs)
        context["approachs"] = Approach.objects.all()
        
        selected_thirdpartys = self.object.第三者点検.values_list('id', flat=True)
        context['selected_thirdpartys'] = list(selected_thirdpartys)
        context["thirdpartys"] = Thirdparty.objects.all()
        
        selected_underconditions = self.object.路下条件.values_list('id', flat=True)
        context['selected_underconditions'] = list(selected_underconditions)
        context["underconditions"] = UnderCondition.objects.all()
        return context

# << ファイルアップロード >>
class UploadView(View):

    def get(self, request, *args, **kwargs):

        context = {}
        context["form"] = TableForm()

        return render(request, "infra/index.html", context)

    def post(self, request, *args, **kwargs):

        form = TableForm(request.POST, request.FILES)

        if form.is_valid():
            print("保存")
            form.save()
        else:
            print(form.errors)

        return redirect("upload")

index   = UploadView.as_view()

# << 名前とアルファベットの登録 >>
def names_list(request, article_pk):
    
    alphabet_list = request.POST.getlist("name_alphabet")
    
    alphabet_list_count = len(alphabet_list)
    for i in range(0, alphabet_list_count, 2):
        dic = {}
        dic["name"] = alphabet_list[i]
        dic["alphabet"] = alphabet_list[i+1]
        dic["article"] = article_pk
        
        form = NameEntryForm(dic)

        if form.is_valid():
            form.save()
        else:
            print(form.errors) # 入力フォームのエラー内容を表示
            
    name_entries = NameEntry.objects.filter(article=article_pk)
    
    return render(request, 'infra/names_list.html', {'article_pk': article_pk, "form": NameEntryForm(), 'name_entries': name_entries})

# << 登録した名前を削除 >>
def delete_name_entry(request, entry_id):
    entry = get_object_or_404(NameEntry, pk=entry_id)
    article_pk = entry.article.pk  # 事前に記事のPKを取得
    if request.method == 'POST':    
        entry.delete()
    name_entries = NameEntry.objects.filter(article=article_pk)
    return render(request, 'infra/names_list.html', {'article_pk': article_pk, "form": NameEntryForm(), 'name_entries': name_entries})

# << 要素番号の登録 >>
def number_list(request, article_pk, pk):
    
    parts_names = PartsName.objects.all().order_by('display_order')  # 順序フィールドで部材名を取り出し並べ替え
    # 同じname属性の値をすべて取り出す
    serial_numbers = request.POST.getlist("serial_number") # ['0101', '0103', '0201', '0203']
    single_numbers = request.POST.getlist("single_number") # ['0101', '0201', '0301', '0401']
    
    new_serial_numbers = []
    serial_numbers_count = len(serial_numbers)
    #    初期値を0 ↓     回数分 ↓           ↓ 2ずつ足す(0101(index:0),0201(index:2))
    for i in range(0, serial_numbers_count, 2):
        new_serial_numbers.append(serial_numbers[i] + "~" + serial_numbers[i+1])
        #                          0101(index:0) ↑          0103(index:1+1) ↑
        #                          0201(index:2) ↑          0203(index:2+1) ↑
    print(new_serial_numbers) # ['0101~0103', '0201~0203']
    
    # 単一の番号、連続の番号 を部材名と紐付けて保存
    for new_serial_number in new_serial_numbers:
        print(new_serial_number)
        if "~" in new_serial_number and len(new_serial_number) >= 5: # 01～02(5文字)
            # new_serial_number = "0101~0205"
            one = new_serial_number.find("~")

            start_number = new_serial_number[:one]
            end_number = new_serial_number[one+1:]

            # 最初の2桁と最後の2桁を取得
            start_prefix = start_number[:2]
            start_suffix = start_number[2:]
            end_prefix = end_number[:2]
            end_suffix = end_number[2:]

            first_elements = []
            # 決められた範囲内の番号を一つずつ追加
            for prefix in range(int(start_prefix), int(end_prefix)+1):
                for suffix in range(int(start_suffix), int(end_suffix)+1):
                    number_items = "{:02d}{:02d}".format(prefix, suffix)
                    dic = {} # forms.pyにも入れないと自動登録ができない
                    dic["number"] = number_items
                    dic["parts_name"] = request.POST.get("parts_name")
                    dic["symbol"] = request.POST.get("symbol")
                    dic["material"] = request.POST.getlist("material")
                    dic["span_number"] = request.POST.get("span_number")
                    dic["main_frame"] = request.POST.get("main_frame") == 'on'
                    dic["infra"] = pk # infraとの紐付け
                    dic["article"] = article_pk
                    print(f"new_serial_number:{number_items}")
                    
                    # 1個ずつバリデーションして保存する
                    form = PartsNumberForm(dic)

                    if form.is_valid():
                        form.save()
                        parts_number = form.save()
                        parts_number.material.set(request.POST.getlist("material"))
                    else:
                        print(form.errors) # 入力フォームのエラー内容を表示
                        
    for single_number in single_numbers:
        if single_number.isdigit():
            dic = {}
            dic["number"] = single_number
            dic["parts_name"] = request.POST.get("parts_name")
            dic["symbol"] = request.POST.get("symbol")
            dic["material"] = request.POST.getlist("material")
            dic["span_number"] = request.POST.get("span_number")
            dic["main_frame"] = request.POST.get("main_frame") == 'on'
            dic["infra"] = pk # infraとの紐付け
            dic["article"] = article_pk 
            print(single_number)

            # 1個ずつバリデーションして保存する
            form    = PartsNumberForm(dic)

            if form.is_valid():
                form.save()
                parts_number = form.save()
                parts_number.material.set(request.POST.getlist("material"))
            else:
                print(form.errors)

    print(f"pk：{pk}、article_pk：{article_pk}")
    create_number_list = PartsNumber.objects.filter(infra=pk)
    print(f"create_number_list：{create_number_list}")
    print("-----------------------------------------")
    print(f"橋梁番号:{pk}") # 橋梁番号:Table object (1)
    print(f"案件番号:{article_pk}") # 案件番号:1
    number_object = Infra.objects.filter(id=pk).first()
    print(f"サイドバーに渡すID：{number_object}")
    for item in create_number_list:
        print(f"Number: {item.number}, Unique ID: {item.unique_id}")

    grouped_parts = defaultdict(list)
    for accordion_list in create_number_list:
        title = f"{accordion_list.parts_name.部材名}（{accordion_list.symbol}）{accordion_list.get_material_list()} {accordion_list.span_number}径間"
        grouped_parts[title].append({
        'number': accordion_list.number,
        'unique_id': accordion_list.unique_id
        })

    return render(request, 'infra/number_entry.html', {'object': number_object, 'article_pk': article_pk, 'pk': pk, "form": PartsNumberForm(), 'create_number_list': create_number_list, 'grouped_parts': grouped_parts.items(), 'parts_names': parts_names})

# << 登録した番号を削除 >>
def delete_number(request, article_pk, pk, unique_id):
    print(f"{article_pk}/{pk}")
    if request.method == 'POST':
        print(f"削除対象：{PartsNumber.objects.filter(infra=pk, article=article_pk)}")
        parts_number = get_object_or_404(PartsNumber, infra=pk, article=article_pk, unique_id=unique_id)
        parts_number.delete()
        return redirect('number-list', article_pk=article_pk, pk=pk)

# << 部材名と記号を紐付けるAjaxリクエスト >>
def get_symbol(request):
    part_id = request.GET.get('part_id')
    try:
        parts_name = PartsName.objects.get(id=part_id)
        return JsonResponse({'symbol': parts_name.記号})
    except PartsName.DoesNotExist:
        return JsonResponse({'error': 'PartsName not found'}, status=404)
    
# 番号表示 TODO:無くても良い
def number_view(request):
    # PartsNumberモデルから1件データを取り出し
    parts_number = PartsNumber.objects.get(id=1)
    # 抽出した数字を文字列として結合
    result = ""
    # 4桁 か 4桁~4桁 のいずれか
    if len(parts_number.number) == 4:
        # 4桁
        result  = parts_number.number
    else:
        # 4桁~4桁
        # ~ で区切る必要がある。 [ "3000","3000" ]
        numbers = parts_number.number.split("~")
        start   = numbers[0]
        end     = numbers[1]
        # 最初の2桁と最後の2桁を取得
        start_prefix = start[:2]
        start_suffix = start[2:]
        end_prefix = end[:2]
        end_suffix = end[2:]

        for prefix in range(int(start_prefix), int(end_prefix)+1):
            for suffix in range(int(start_suffix), int(end_suffix)+1):
                result += "{:02d}{:02d}\n".format(prefix, suffix)

    print(result)

def match_s3_objects_with_prefix(bucket_name, prefix, pattern):
    s3 = boto3.client('s3')
    # プレフィックス(特定のフォルダ)を指定して、オブジェクトをリスト
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
    if 'Contents' not in response:
        return []
    # パターンに基づいてオブジェクトをフィルタリング
    matched_keys = [obj['Key'] for obj in response['Contents'] if fnmatch.fnmatch(obj['Key'], pattern)]
    return matched_keys


# << 損傷写真帳の作成 >>
start1 = time.time()
def bridge_table(request, article_pk, pk): # idの紐付け infra/bridge_table.htmlに表示
    
    context = {}
    # プロジェクトのメディアディレクトリからdxfファイルまでの相対パス
    # URL：article/<int:article_pk>/infra/<int:pk>/bridge-table/

    # 指定したInfraに紐づく Tableを取り出す
    article = Article.objects.filter(id=article_pk).first()
    infra = Infra.objects.filter(id=pk).first()
    table = Table.objects.filter(infra=pk).first()
    # << 案件名とファイル名を連結してdxfファイルのURLを取得する >>
    # AWSクライアントを作成
    s3 = boto3.client('s3')
    
    bucket_name = 'infraprotect'
    print(bucket_name)
    folder_name = article.案件名+"/"
    print(folder_name)
    pattern = f'*{infra.title}*/{infra.title}.dxf'
    print(pattern)

    # 該当するオブジェクトを取得
    matched_objects = match_s3_objects_with_prefix(bucket_name, folder_name, pattern)

    if matched_objects:
        print(f"該当オブジェクト：{matched_objects}")
    else:
        print("ファイルが見つかりません")

    # 結果を表示
    for obj_key in matched_objects:
        encode_dxf_filename = f"https://{bucket_name}.s3.ap-northeast-1.amazonaws.com/{obj_key}"
    
    dxf_filename = urllib.parse.quote(encode_dxf_filename, safe='/:') # スラッシュとコロン以外をエンコード

    print(f"dxfファイルのデコードURLは：{encode_dxf_filename}")
    print(f"dxfファイルの絶対URLは：{dxf_filename}")
    
    # bridge_tableのボタンを押したときのアクション
    second_search_title_text = "損傷図"
    # << 辞書型として、全径間を1つの多重リストに格納 >>
    max_search_title_text = infra.径間数
    print(f"最大径間数：{max_search_title_text}")
    database_sorted_items = []  # 結果をまとめるリスト
    
    for search_title_text_with_suffix in range(1, max_search_title_text + 1):
        search_title_text = f"{search_title_text_with_suffix}径間"
        
        print(search_title_text)
        #                                                                                   1径間                  損傷図
        sub_database_sorted_items = create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text)

        for item in sub_database_sorted_items:
            item['search'] = search_title_text
            database_sorted_items.append(item)

    """辞書型の多重リストをデータベースに登録"""
    # << ['']を外してフラットにする >>
    def flatten(value):
        def _flatten(nested_list):
            if isinstance(nested_list, list):
                for item in nested_list:
                    yield from _flatten(item)
            else:
                yield nested_list
        
        return ', '.join(_flatten(value))

    # << joinキーを変換 >>
    def join_to_result_string(join):
        result_parts = []
        for item in join:
            parts_name = item['parts_name'][0]
            damage_names = item['damage_name']
            formatted_damage_names = '/'.join(damage_names)
            result_parts.append(f"{parts_name} : {formatted_damage_names}")
        return ', '.join(result_parts)

    # << 写真のキーを変換 >>
    def simple_flatten(value):
        return ', '.join(map(str, value)) if isinstance(value, list) else value
    
    # <<正規表現で4桁以上の番号を取得>>
    def extract_number(text):
        pattern = r'\d{4,}' # 4文字以上の連続する数字
        matches = re.findall(pattern, text)
        return matches
    
    picture_counter = 1
    index_counter = 0
    picture_number_box = []

    for damage_data in database_sorted_items:
        # 元の辞書から 'picture_number' の値を取得
        #             　辞書型 ↓           ↓ キーの名前      ↓ 存在しない場合、デフォルト値として空白を返す
        picture_number = damage_data.get('picture_number', '')
        # 正規表現で数字のみを抽出
        if picture_number:
            # 数字のみを抽出
            before_numbers_only = re.findall(r'\d+', str(picture_number)) # ['2']  ['2','3']
            #print(f"リスト型番号:{before_numbers_only}")
            #print(f"{index_counter}  どっちが大きい　{len(before_numbers_only)}")
            # before_numbers_onlyの各元素で別の処理を行う場合
            # カウンターに基づいて処理を行う
            if index_counter == 0:
                picture_number_box = []
            if len(before_numbers_only) > 1:
                for number in before_numbers_only:
                    #print(f"{index_counter}番目の要素: {number}")
                    picture_number_box.append(number)
                    index_counter += 1
                index_counter = 0
                #print(picture_number_box)
            else:
                picture_number_box = []
                index_counter = 0
                numbers_only = before_numbers_only[index_counter]  # カウンターに対応する数字を取得
                #print(f"オンリーナンバーズ（抽出後）: {numbers_only}")
                picture_number_box.append(numbers_only)
        else:
            numbers_only = None

        damage_coordinate = damage_data.get('damage_coordinate', [None, None])
        damage_coordinate_x = damage_coordinate[0] if damage_coordinate else None
        damage_coordinate_y = damage_coordinate[1] if damage_coordinate else None

        picture_coordinate = damage_data.get('picture_coordinate', [None, None])
        picture_coordinate_x = picture_coordinate[0] if picture_coordinate else None
        picture_coordinate_y = picture_coordinate[1] if picture_coordinate else None

        #parts_list = flatten(damage_data.get('parts_name', ''))
        #damage_list = flatten(damage_data.get('damage_name', ''))

        names = damage_data.get('parts_name', '')
        damages = damage_data.get('damage_name', '')
        #print(f"names:{names}")
        print(f"damages:{damages}")
        
        split_names = []

        for item in names:
            split_items = []
            for split in item:
                if "～" in split:
                    one = split.find("～")
                    start_number = ''.join(extract_number(split[:one])) # 0101
                    end_number = ''.join(extract_number(split[one+1:])) # 0204

                    # 最初の2桁と最後の2桁を取得
                    start_prefix = start_number[:2] # 01
                    start_suffix = start_number[2:] # 01
                    end_prefix = end_number[:2] # 01
                    end_suffix = end_number[2:] # 03
                    
                    part_name = split[:one].replace(start_number, '')
                
                    for prefix in range(int(start_prefix), int(end_prefix)+1):
                        for suffix in range(int(start_suffix), int(end_suffix)+1):
                            number_items = "{:02d}{:02d}".format(prefix, suffix)
                            split_items.append(part_name + number_items)
                else:
                    split_items.append(split)
            split_names.append(split_items)
        
        join = join_to_result_string(damage_data.get('join', ''))
        this_time_picture = simple_flatten(damage_data.get('this_time_picture', ''))
        last_time_picture = simple_flatten(damage_data.get('last_time_picture', ''))
        textarea_content = damage_data.get('textarea_content', '')
        span_number = damage_data.get('search', '')
        #print("----------------------------------")
        #print(f"damage_data:{damage_data}")
        #print(f"join:{join}")
        #print(f"this_time_picture:{this_time_picture}")
        name_length = len(split_names)
        damage_length = len(damages)
        
        # 多重リストかどうかを判定する関数
        def is_multi_list(lst):
            return any(isinstance(i, list) for i in lst)
        
        def process_names(names):
            """
            与えられたnamesを処理し、適切な部分を返す関数
            所見用にparts_splitに格納
            """
            parts_left = ["主桁", "PC定着部"]# 左の数字
            parts_right = ["横桁", "橋脚[柱部・壁部]", "橋脚[梁部]", "橋脚[隅角部・接合部]", "橋台[胸壁]", "橋台[竪壁]", "橋台[翼壁]"]# 右の数字
            parts_zero = ["床版"]# 00になる場合

            # namesから部品名（parts）と数字を抽出
            space = names.find(" ")
            parts = names[:space]  # 部品名
            number = ''.join(extract_number(names))  # 数字
            parts_join = names.replace(number, '') # 符号部分を取得

            # 必要な部分の数字を抽出するロジック
            split_number = ''

            if parts in parts_zero:
                split_number = '00'
            elif len(number) == 4 or int(number[2:]) >= 100:
                if parts in parts_left:
                    split_number = number[:2]
                elif parts in parts_right:
                    split_number = number[2:]
                else:
                    split_number = '00'
            else:
                if parts in parts_left:
                    split_number = number[:3]
                elif parts in parts_right:
                    split_number = number[3:]
                else:
                    split_number = '00'

            result = parts_join + split_number  # 結果を組み立てる
            return result
            # 共通のフィールドを辞書に格納
        infra = Infra.objects.filter(id=pk).first()
        article = infra.article
        table = Table.objects.filter(infra=infra.id, article=article.id).first()
        #print(table) # 旗揚げチェック：お試し（infra/table/dxf/121_2径間番号違い.dxf）
        
        # << 管理サイトに登録するコード（損傷写真帳） >>
        split_items_table = is_multi_list(split_names)
        damages_items_table = is_multi_list(damages)
        
        if not split_items_table and not damages_items_table and name_length == 1: # 部材名が1つの場合
            picture_number_index = 0 # 写真番号は0から始める
            for single_damage in damages: 
                parts_name = names[0]
                pattern = r"(\d+)$"
                # parts_nameからパターンにマッチする部分を検索
                match = re.search(pattern, parts_name)
                if match:
                    four_numbers = match.group(1)
                else:
                    four_numbers = "00"
                damage_name = flatten(single_damage)
                # print(f"parts_name1:{parts_name}") # 主桁 Mg0101
                # print(f"damage_name1:{damage_name}") # ㉓変形・欠損-c
                parts_split = process_names(flatten(parts_name))
                # print(f"this_time:{this_time_picture}")
                list_in_picture = this_time_picture.split(",")
                for single_picture in list_in_picture:
                    update_fields = {
                        'parts_name': parts_name,
                        'four_numbers': four_numbers,
                        'damage_name': damage_name,
                        'parts_split': parts_split,
                        'join': join,
                        'this_time_picture': single_picture,
                        'last_time_picture': last_time_picture,
                        'textarea_content': textarea_content,
                        'damage_coordinate_x': damage_coordinate_x,
                        'damage_coordinate_y': damage_coordinate_y,
                        'picture_coordinate_x': picture_coordinate_x,
                        'picture_coordinate_y': picture_coordinate_y,
                        'span_number': span_number,
                        'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                        'infra': infra,
                        'article': article,
                        'table': table
                    }
                    # print(f"update_fields：{update_fields}")
                #print(f"径間番号:{span_number}")
                # if FullReportData.objects.filter(join=join, this_time_picture=this_time_picture, span_number=span_number, table=table, damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y):
                    # update_fields['this_time_picture'] = ""

                    if single_picture:
                        images = [single_picture]
                        update_fields['picture_number'] = picture_counter
                        picture_counter += 1
                        # BridgePictureモデルに保存
                        if numbers_only is not None and numbers_only != '':
                            # print(f"images：{images}") # ['写真パス1', '写真パス2']
                            for absolute_image_path in images:
                                # print(f"absolute_image_path：{absolute_image_path}") # 写真パス1（別データとして、写真パス2）
                                try:
                                # with open(absolute_image_path, 'rb') as image_file:
                                    #print(f"保存前：{numbers_only}")
                                    #print(f"オンリーナンバーズ（抽出後）: {picture_number_box}")
                                    if picture_number_index < len(picture_number_box):
                                        current_picture_number = picture_number_box[picture_number_index]
                                    else:
                                        current_picture_number = None
                                    # print(f"保存後：{current_picture_number}")
                                    #print("---------")
                                    if current_picture_number is None: # 写真がない
                                        join_picture_damage_name = BridgePicture.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                        # print(f"join_picture_damage_name：{join_picture_damage_name}")
                                        if join_picture_damage_name.first():
                                            for picture in join_picture_damage_name:
                                                #print(picture)
                                                if picture.damage_name:
                                                    # print(f"損傷名：{picture.damage_name}") # 損傷名
                                                    edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                                    new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                                    # 末尾のハイフン+任意の1文字を削除
                                                    damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                                    picture.memo = f"{picture.memo} / {parts_split},{damage_name}"
                                                else:
                                                    picture.memo = f"{parts_split},{damage_name}"
                                                picture.save()
                                            #print(join_picture_damage_name)
                                        #print("picture_number_boxのインデックスが範囲外です")
                                        continue
                                    # 「スペース + 2文字以上のアルファベット + 2文字以上の数字」にマッチする部分を捉える
                                    pattern = r'\s+[A-Za-z]{2,}[0-9]{2,}'
                                    # マッチした部分からアルファベット部分だけを削除するための関数を定義
                                    def remove_alphabets(match):
                                        # マッチした文字列からアルファベット部分を削除
                                        return re.sub(r'[A-Za-z]+', '', match.group())
                                            
                                    joined_picture_damage_name = FullReportData.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                    print(f"joined_picture_damage_name：{joined_picture_damage_name}")
                                    
                                    loop_change = True
                                    for full_damaged_name in joined_picture_damage_name:
                                        if loop_change:
                                            print(full_damaged_name.join)
                                            loop_change = False
                                            
                                    # re.subでパターンにマッチする部分を編集
                                    edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                    new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                    damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                    # 写真の重複チェック(写真番号が同じ、損傷座標・def座標が同じ、径間番号・dxfファイル名・案件名・橋梁名が同じ)
                                    existing_picture = BridgePicture.objects.filter(
                                        picture_number=current_picture_number,
                                        damage_coordinate_x=damage_coordinate_x,
                                        damage_coordinate_y=damage_coordinate_y,
                                        picture_coordinate_x=picture_coordinate_x,
                                        picture_coordinate_y=picture_coordinate_y,
                                        span_number=span_number,
                                        table=table,
                                        article=article,
                                        infra=infra
                                    ).first()
                                    
                                    if existing_picture is None:
                                        bridge_picture = BridgePicture(
                                            image=absolute_image_path, 
                                            picture_number=current_picture_number,
                                            damage_name=damage_name,
                                            parts_split=edited_result_parts_name,
                                            memo=process_inspection_data(full_damaged_name.join),
                                            damage_coordinate_x=damage_coordinate_x,
                                            damage_coordinate_y=damage_coordinate_y,
                                            picture_coordinate_x=picture_coordinate_x,
                                            picture_coordinate_y=picture_coordinate_y,
                                            span_number=span_number,
                                            table=table,
                                            article=article,
                                            infra=infra
                                        )
                                        bridge_picture.save()
                                        picture_number_index += 1
                                except FileNotFoundError:
                                    print(f"ファイルが見つかりません: {absolute_image_path}")
           
                    else:
                        update_fields['picture_number'] = ""
                    
                report_data_exists = FullReportData.objects.filter(**update_fields).exists()
                if report_data_exists:
                    print("データが存在しています。")
                else:
                    try:
                        damage_obj, created = FullReportData.objects.update_or_create(**update_fields)
                        damage_obj.save()
                    except IntegrityError:
                        print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")
                    
                    
        elif not split_items_table and not damages_items_table and name_length >= 2: # 部材名が2つ以上の場合
            picture_number_index = 0
            if damage_length == 1: # かつ損傷名が1つの場合
                for parts_name in split_names:
                    pattern = r"(\d+)$"
                    # parts_nameからパターンにマッチする部分を検索
                    match = re.search(pattern, parts_name)
                    if match:
                        four_numbers = match.group(1)
                    else:
                        four_numbers = "00"
                    damage_name = flatten(damages[0])
                    # print(f"parts_name2:{parts_name}")
                    # print(f"damage_name2:{damage_name}")
                    parts_split = process_names(flatten(parts_name))
                    # print(f"this_time:{this_time_picture}")
                    list_in_picture = this_time_picture.split(",")
                    for single_picture in list_in_picture:
                        update_fields = {
                            'parts_name': parts_name,
                            'four_numbers': four_numbers,
                            'damage_name': damage_name,
                            'parts_split': parts_split,
                            'join': join,
                            'this_time_picture': single_picture,
                            'last_time_picture': last_time_picture,
                            'textarea_content': textarea_content,
                            'damage_coordinate_x': damage_coordinate_x,
                            'damage_coordinate_y': damage_coordinate_y,
                            'picture_coordinate_x': picture_coordinate_x,
                            'picture_coordinate_y': picture_coordinate_y,
                            'span_number': span_number,
                            'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                            'infra': infra,
                            'article': article,
                            'table': table
                        }         
                    #print(f"径間番号:{span_number}")                 
                    # if FullReportData.objects.filter(join=join, this_time_picture=this_time_picture, span_number=span_number, table=table, damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y):
                        # update_fields['this_time_picture'] = ""
                        # update_fields['picture_number'] = ""
                        
                        if single_picture:
                            images = [single_picture]
                            update_fields['picture_number'] = picture_counter
                            picture_counter += 1
                            # BridgePictureモデルに保存
                            if numbers_only is not None and numbers_only != '':
                                # print(f"images：{images}")
                                for absolute_image_path in images:
                                    # print(f"absolute_image_path：{absolute_image_path}")
                                    try:
                                    # with open(absolute_image_path, 'rb') as image_file:
                                        #print(f"保存前：{numbers_only}")
                                        #print(f"オンリーナンバーズ（抽出後）: {picture_number_box}")
                                        if picture_number_index < len(picture_number_box):
                                            current_picture_number = picture_number_box[picture_number_index]
                                        else:
                                            current_picture_number = None
                                        # print(f"保存後：{current_picture_number}")
                                        #print("---------")
                                        if current_picture_number is None: # 写真がない
                                            join_picture_damage_name = BridgePicture.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                            
                                            # print(f"join_picture_damage_name：{join_picture_damage_name}")
                                            if join_picture_damage_name.first():
                                                for picture in join_picture_damage_name:
                                                    if picture.damage_name:
                                                        # print(f"損傷名：{picture.damage_name}") # 損傷名
                                                        edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                                        new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                                        # 末尾のハイフン+任意の1文字を削除
                                                        damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                                        picture.memo = f"{picture.memo} / {parts_split},{damage_name}"
                                                    else:
                                                        picture.memo = f"{parts_split},{damage_name}"
                                                    picture.save()
                                                #print(join_picture_damage_name)
                                            #print("picture_number_boxのインデックスが範囲外です")
                                            continue
                                        # 「スペース + 2文字以上のアルファベット + 2文字以上の数字」にマッチする部分を捉える
                                        pattern = r'\s+[A-Za-z]{2,}[0-9]{2,}'
                                        # マッチした部分からアルファベット部分だけを削除するための関数を定義
                                        def remove_alphabets(match):
                                            # マッチした文字列からアルファベット部分を削除
                                            return re.sub(r'[A-Za-z]+', '', match.group())
                                            
                                        joined_picture_damage_name = FullReportData.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                        # print(f"joined_picture_damage_name：{joined_picture_damage_name}")
                                        
                                        loop_change = True
                                        for full_damaged_name in joined_picture_damage_name:
                                            if loop_change:
                                                print(full_damaged_name.join)
                                                loop_change = False
                                                
                                        # re.subでパターンにマッチする部分を編集
                                        edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                        new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                        damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                        existing_picture = BridgePicture.objects.filter(
                                            picture_number=current_picture_number,
                                            damage_coordinate_x=damage_coordinate_x,
                                            damage_coordinate_y=damage_coordinate_y,
                                            picture_coordinate_x=picture_coordinate_x,
                                            picture_coordinate_y=picture_coordinate_y,
                                            span_number=span_number,
                                            table=table,
                                            article=article,
                                            infra=infra
                                        ).first()
                                        
                                        if existing_picture is None:
                                            bridge_picture = BridgePicture(
                                                image=absolute_image_path,
                                                picture_number=current_picture_number,
                                                damage_name=damage_name,
                                                parts_split=edited_result_parts_name,
                                                memo=process_inspection_data(full_damaged_name.join),
                                                damage_coordinate_x=damage_coordinate_x,
                                                damage_coordinate_y=damage_coordinate_y,
                                                picture_coordinate_x=picture_coordinate_x,
                                                picture_coordinate_y=picture_coordinate_y,
                                                span_number=span_number,
                                                table=table,
                                                article=article,
                                                infra=infra
                                            )
                                            bridge_picture.save()
                                            picture_number_index += 1
                                    except FileNotFoundError:
                                        print(f"ファイルが見つかりません: {absolute_image_path}")
             
                        else:
                            update_fields['picture_number'] = ""
                    
                    report_data_exists = FullReportData.objects.filter(**update_fields).exists()
                    if report_data_exists:
                        print("データが存在しています。")
                    else:
                        try:
                            damage_obj, created = FullReportData.objects.update_or_create(**update_fields)
                            damage_obj.save()
                        except IntegrityError:
                            print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")
                        
            elif not split_items_table and not damages_items_table and damage_length >= 2: # かつ損傷名が2つ以上の場合
                picture_number_index = 0
                for name in split_names:
                    for damage in damages:
                        parts_name = name
                        pattern = r"(\d+)$"
                        # parts_nameからパターンにマッチする部分を検索
                        match = re.search(pattern, parts_name)
                        if match:
                            four_numbers = match.group(1)
                        else:
                            four_numbers = "00"
                        damage_name = flatten(damage)
                        # print(f"parts_name3:{parts_name}")
                        # print(f"damage_name3:{damage_name}")
                        parts_split = process_names(flatten(parts_name))
                        # print(f"this_time:{this_time_picture}")
                        list_in_picture = this_time_picture.split(",")
                        for single_picture in list_in_picture:
                            update_fields = {
                                'parts_name': parts_name,
                                'four_numbers': four_numbers,
                                'damage_name': damage_name,
                                'parts_split': parts_split,
                                'join': join,
                                'this_time_picture': single_picture,
                                'last_time_picture': last_time_picture,
                                'textarea_content': textarea_content,
                                'damage_coordinate_x': damage_coordinate_x,
                                'damage_coordinate_y': damage_coordinate_y,
                                'picture_coordinate_x': picture_coordinate_x,
                                'picture_coordinate_y': picture_coordinate_y,
                                'span_number': span_number,
                                'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                                'infra': infra,
                                'article': article,
                                'table': table
                            }
                        #print(f"径間番号:{span_number}")
                        # if FullReportData.objects.filter(join=join, this_time_picture=this_time_picture, span_number=span_number, table=table, damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y):
                            # update_fields['this_time_picture'] = ""
                            # update_fields['picture_number'] = ""
                            
                            if single_picture:
                                images = [single_picture]
                                update_fields['picture_number'] = picture_counter
                                picture_counter += 1
                                # BridgePictureモデルに保存
                                if numbers_only is not None and numbers_only != '':
                                    # print(f"images：{images}")
                                    for absolute_image_path in images:
                                        # print(f"absolute_image_path：{absolute_image_path}")
                                        try:
                                        # with open(absolute_image_path, 'rb') as image_file:
                                            #print(f"保存前：{numbers_only}")
                                            #print(f"オンリーナンバーズ（抽出後）: {picture_number_box}")
                                            if picture_number_index < len(picture_number_box):
                                                current_picture_number = picture_number_box[picture_number_index]
                                            else:
                                                current_picture_number = None
                                            # print(f"保存後：{current_picture_number}")
                                            # print("---------")
                                            if current_picture_number is None: # 写真がない
                                                join_picture_damage_name = BridgePicture.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                                # print(f"join_picture_damage_name：{join_picture_damage_name}")
                                                    
                                                if join_picture_damage_name.first():
                                                    for picture in join_picture_damage_name:
                                                        #print(picture)
                                                        if picture.damage_name:
                                                            # print(f"損傷名：{picture.damage_name}") # 損傷名
                                                            edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                                            new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                                            # 末尾のハイフン+任意の1文字を削除
                                                            damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                                            picture.memo = f"{picture.memo} / {parts_split},{damage_name}"
                                                        else:
                                                            picture.memo = f"{parts_split},{damage_name}"
                                                        picture.save()
                                                    
                                                #print("picture_number_boxのインデックスが範囲外です")
                                                continue
                                            # 「スペース + 2文字以上のアルファベット + 2文字以上の数字」にマッチする部分を捉える
                                            pattern = r'\s+[A-Za-z]{2,}[0-9]{2,}'
                                            # マッチした部分からアルファベット部分だけを削除するための関数を定義
                                            def remove_alphabets(match):
                                                # マッチした文字列からアルファベット部分を削除
                                                return re.sub(r'[A-Za-z]+', '', match.group())
                                            
                                            joined_picture_damage_name = FullReportData.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                            # print(f"joined_picture_damage_name：{joined_picture_damage_name}")
                                            
                                            loop_change = True
                                            for full_damaged_name in joined_picture_damage_name:
                                                if loop_change:
                                                    # print(full_damaged_name.join)
                                                    loop_change = False
                                                    
                                            # re.subでパターンにマッチする部分を編集
                                            edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                            new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', damage_name)
                                            damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                            existing_picture = BridgePicture.objects.filter(
                                                picture_number=current_picture_number,
                                                damage_coordinate_x=damage_coordinate_x,
                                                damage_coordinate_y=damage_coordinate_y,
                                                picture_coordinate_x=picture_coordinate_x,
                                                picture_coordinate_y=picture_coordinate_y,
                                                span_number=span_number,
                                                table=table,
                                                article=article,
                                                infra=infra
                                            ).first()
                                            
                                            if existing_picture is None:
                                                bridge_picture = BridgePicture(
                                                    image=absolute_image_path,
                                                    picture_number=current_picture_number,
                                                    damage_name=damage_name,
                                                    parts_split=edited_result_parts_name,
                                                    memo=process_inspection_data(full_damaged_name.join),
                                                    damage_coordinate_x=damage_coordinate_x,
                                                    damage_coordinate_y=damage_coordinate_y,
                                                    picture_coordinate_x=picture_coordinate_x,
                                                    picture_coordinate_y=picture_coordinate_y,
                                                    span_number=span_number,
                                                    table=table,
                                                    article=article,
                                                    infra=infra
                                                )
                                                bridge_picture.save()
                                                picture_number_index += 1
                                        except FileNotFoundError:
                                            print(f"ファイルが見つかりません: {absolute_image_path}")
                      
                            else:
                                update_fields['picture_number'] = ""
                            
                        report_data_exists = FullReportData.objects.filter(**update_fields).exists()
                        if report_data_exists:
                            print("データが存在しています。")
                        else:
                            try:
                                damage_obj, created = FullReportData.objects.update_or_create(**update_fields)
                                damage_obj.save()
                            except IntegrityError:
                                print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")
                                 
        else: # 多重リストの場合
            picture_number_index = 0
            for i in range(name_length):
                for name in split_names[i]:
                    for damage in damages[i]:
                        parts_name = name
                        
                        pattern = r"(\d+)$"
                        match = re.search(pattern, parts_name)
                        if match:
                            four_numbers = match.group(1)
                        else:
                            four_numbers = "00"
                        original_damage_name = flatten(damage)
                        # print(f"parts_name4:{parts_name}")
                        # print(f"damage_name4:{original_damage_name}")
                        parts_split = process_names(flatten(parts_name))
                        # print(f"this_time:{this_time_picture}")
                        list_in_picture = this_time_picture.split(",")
                        picture_number_index = 0 # インデックス番号のリセット
                        for single_picture in list_in_picture:
                            update_fields = {
                                'parts_name': parts_name,
                                'four_numbers': four_numbers,
                                'damage_name': str(original_damage_name),
                                'parts_split': parts_split,
                                'join': join,
                                'this_time_picture': single_picture.strip(),
                                'last_time_picture': last_time_picture,
                                'textarea_content': textarea_content,
                                'damage_coordinate_x': damage_coordinate_x,
                                'damage_coordinate_y': damage_coordinate_y,
                                'picture_coordinate_x': picture_coordinate_x,
                                'picture_coordinate_y': picture_coordinate_y,
                                'span_number': span_number,
                                'special_links': '/'.join([str(parts_split), str(original_damage_name), str(span_number)]),
                                'infra': infra,
                                'article': article,
                                'table': table
                            }
                            # print(f"update_fields：{update_fields}")
                            
                            # この部分で今生成したupdate_fieldsでデータを保存または更新します
                            report_data_exists = FullReportData.objects.filter(**update_fields).exists()
                            if report_data_exists:
                                print("データが既に存在しています。")
                            else:
                                try:
                                    damage_obj, created = FullReportData.objects.update_or_create(**update_fields)
                                    damage_obj.save()
                                except IntegrityError:
                                    print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")
                            
                            if single_picture:
                                images = [single_picture]
                                update_fields['picture_number'] = picture_counter
                                picture_counter += 1
                                # BridgePictureモデルに保存
                                if numbers_only is not None and numbers_only != '':
                                    # print(f"images：{images}")
                                    for absolute_image_path in images:
                                        # print(f"absolute_image_path：{absolute_image_path}")
                                        try:
                                        # with open(absolute_image_path, 'rb') as image_file:
                                            #print(f"保存前：{numbers_only}")
                                            #print(f"オンリーナンバーズ（抽出後）: {picture_number_box}")
                                            
                                            # print(f"picture_number_index：{picture_number_index}") # 0/1
                                            # print(f"picture_number_box：{picture_number_box}") # ['10', '11']
                                            # print(f"len(picture_number_box)：{len(picture_number_box)}") # 2
                                            # print(f"picture_number_box[picture_number_index]：{picture_number_box[picture_number_index]}") # 10/11
                                            if picture_number_index < len(picture_number_box):
                                                current_picture_number = picture_number_box[picture_number_index]
                                            else:
                                                current_picture_number = None
                                                
                                            # print(f"保存後：{current_picture_number}")
                                            #print("---------")
                                            if current_picture_number is None: # 写真がない
                                                join_picture_damage_name = BridgePicture.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                                # print(f"join_picture_damage_name：{join_picture_damage_name}")
                                                if join_picture_damage_name.first():
                                                    for picture in join_picture_damage_name:
                                                        # print(picture)
                                                        if picture.damage_name:
                                                            # print(f"損傷名：{picture.damage_name}") # 損傷名
                                                            edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                                            new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', original_damage_name)
                                                            # 末尾のハイフン+任意の1文字を削除
                                                            damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                                            picture.memo = f"{picture.memo} / {parts_split},{damage_name}"
                                                        else:
                                                            picture.memo = f"{parts_split},{damage_name}"
                                                        picture.save()
                                                    #print(join_picture_damage_name)
                                                #print("picture_number_boxのインデックスが範囲外です")
                                                continue
                                            # 「スペース + 2文字以上のアルファベット + 2文字以上の数字」にマッチする部分を捉える
                                            pattern = r'\s+[A-Za-z]{2,}[0-9]{2,}'
                                            # マッチした部分からアルファベット部分だけを削除するための関数を定義
                                            def remove_alphabets(match):
                                                # マッチした文字列からアルファベット部分を削除
                                                return re.sub(r'[A-Za-z]+', '', match.group())
                                            joined_picture_damage_name = FullReportData.objects.filter(damage_coordinate_x=damage_coordinate_x, damage_coordinate_y=damage_coordinate_y, table=table, infra=infra, article=article)
                                            # print(f"joined_picture_damage_name：{joined_picture_damage_name}")
                                            
                                            loop_change = True
                                            for full_damaged_name in joined_picture_damage_name:
                                                if loop_change:
                                                    print(full_damaged_name.join)
                                                    loop_change = False

                                            # re.subでパターンにマッチする部分を編集
                                            edited_result_parts_name = re.sub(pattern, remove_alphabets, parts_split)
                                            # print(f"original_damage_name:{original_damage_name}")
                                            new_damage_name = re.sub(r'^[\u2460-\u2473\u3251-\u3256]', '', original_damage_name)
                                            # print(f"new_damage_name：{new_damage_name}")
                                            # damage_name = re.sub(r'-.{1}$', '', new_damage_name)
                                            # print(damage_name)
                                            existing_picture = BridgePicture.objects.filter(
                                                image=absolute_image_path,
                                                picture_number=current_picture_number,
                                                damage_coordinate_x=damage_coordinate_x,
                                                damage_coordinate_y=damage_coordinate_y,
                                                picture_coordinate_x=picture_coordinate_x,
                                                picture_coordinate_y=picture_coordinate_y,
                                                span_number=span_number,
                                                table=table,
                                                article=article,
                                                infra=infra
                                            ).first()
                                            # print(f"absolute_image_path：{absolute_image_path}") # 写真URL
                                            # print(f"current_picture_number：{current_picture_number}") # 10
                                            # print(f"data：{existing_picture}") # BridgePicture object (946)
                                            
                                            if existing_picture is None:
                                                try:
                                                    bridge_picture = BridgePicture(
                                                        image=absolute_image_path, 
                                                        picture_number=current_picture_number,
                                                        damage_name=original_damage_name,
                                                        parts_split=edited_result_parts_name,
                                                        memo=process_inspection_data(full_damaged_name.join),
                                                        damage_coordinate_x=damage_coordinate_x,
                                                        damage_coordinate_y=damage_coordinate_y,
                                                        picture_coordinate_x=picture_coordinate_x,
                                                        picture_coordinate_y=picture_coordinate_y,
                                                        span_number=span_number,
                                                        table=table,
                                                        article=article,
                                                        infra=infra
                                                    )
                                                    bridge_picture.save()
                                                except:
                                                    index = images.index(absolute_image_path)
                                                    if index +1 >= len(images):
                                                        index = len(images) -1
                                                    bridge_picture = BridgePicture(
                                                        image=images[index], 
                                                        picture_number=current_picture_number,
                                                        damage_name=original_damage_name,
                                                        parts_split=edited_result_parts_name,
                                                        memo=process_inspection_data(full_damaged_name.join),
                                                        damage_coordinate_x=damage_coordinate_x,
                                                        damage_coordinate_y=damage_coordinate_y,
                                                        picture_coordinate_x=picture_coordinate_x,
                                                        picture_coordinate_y=picture_coordinate_y,
                                                        span_number=span_number,
                                                        table=table,
                                                        article=article,
                                                        infra=infra
                                                    )
                                                    # print(f"image：{images[index]}")
                                                    try:
                                                        bridge_picture.save()
                                                    except:
                                                        print("保存失敗")
                                                # print(f"bridge_picture：{bridge_picture}")
                                                picture_number_index += 1
                                        except FileNotFoundError:
                                            print(f"ファイルが見つかりません: {absolute_image_path}")
                                    
                            else:
                                update_fields['picture_number'] = ""
                        
                        report_data_exists = FullReportData.objects.filter(**update_fields).exists()
                        if report_data_exists:
                            print("データが存在しています。")
                        else:
                            try:
                                damage_obj, created = FullReportData.objects.update_or_create(**update_fields)
                                damage_obj.save()
                            except IntegrityError:
                                print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")
                                
    """辞書型の多重リストをデータベースに登録(ここまで)"""

    if "search_title_text" in request.GET:
        # request.GET：検索URL（http://127.0.0.1:8000/article/1/infra/bridge_table/?search_title_text=1径間） 
        search_title_text = request.GET["search_title_text"]
        # 検索URL内のsearch_title_textの値（1径間）を取得する
    else:
        search_title_text = "1径間" # 検索URLにsearch_title_textがない場合
    second_search_title_text = "損傷図"

    bridges = FullReportData.objects.filter(infra=pk, span_number=search_title_text) # 径間で絞り込み
    # parts_name のカスタム順序リスト
    parts_order = ['主桁', '横桁', '床版', 'PC定着部', '橋台[胸壁]', '橋台[竪壁]', '橋台[翼壁]', '支承本体', '沓座モルタル', '防護柵', '地覆', '伸縮装置', '舗装', '排水ます', '排水管']
    damage_order = ['①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩', '⑪', '⑫', '⑬', '⑭', '⑮', '⑯', '⑰', '⑱', '⑲', '⑳', '㉑', '㉒', '㉓', '㉔', '㉕', '㉖']

    grouped_data = []
    for key, group in groupby(bridges, key=attrgetter('join', 'damage_coordinate_x', 'damage_coordinate_y')):
        grouped_data.append(list(group))

    photo_grouped_data = []
    for pic_key, pic_group in groupby(bridges, key=attrgetter('this_time_picture', 'span_number')):
        photo_grouped_data.append(list(pic_group))
    
    buttons_count = int(table.infra.径間数) # 数値として扱う
    buttons = list(range(1, buttons_count + 1)) # For loopのためのリストを作成
    
    # range(一連の整数を作成):range(1からスタート, ストップ引数3 = 2 + 1) → [1, 2](ストップ引数は含まれない)
    print(buttons)
    
    print(f"ボタン:{Table.objects.filter(infra=pk)}")# ボタン:<QuerySet [<Table: Table object (1)>]>
        # クエリセットを使って対象のオブジェクトを取得
    table_object = Infra.objects.filter(id=pk).first()    
    print(f"橋梁番号:{table_object}")# ボタン:Table object (1)
    print(f"橋梁番号:{table_object.id}")
    article_pk = infra.article.id
    print(f"案件番号:{article_pk}") # 案件番号:1
    
    picture_data = [] # ここで毎回初期化されます
    for data in bridges:
        # クエリセットでフィルタリング
        matches = BridgePicture.objects.filter(
            picture_coordinate_x=data.picture_coordinate_x,
            picture_coordinate_y=data.picture_coordinate_y,
            span_number=data.span_number,
            table=data.table,
            infra=data.infra,
            article=data.article
        ).distinct()
        # picture_data.append({"full_report": data, "matches": matches})
        # matches の各要素に対して個別に処理

        # matches に QuerySet を含めるために新たに作成
        match_details = [
            {"id": match.picture_number, "other_field": match.image}
            for match in matches
        ]
            
        picture_data.append({"full_report": data, "matches": matches, "picture": match_details})

    # matchesのキーに基づき、picturesを交互に選択してユニークに
    matches_seen = {}
    for entry in picture_data:
        matches_key = tuple(entry['matches'].values_list('id', flat=True))
        if matches_key not in matches_seen:
            matches_seen[matches_key] = 0
        
        if len(entry['picture']) != 0:
            picture_index = matches_seen[matches_key] % len(entry['picture'])
            entry['picture'] = [entry['picture'][picture_index]]
            # 次に進める
            matches_seen[matches_key] += 1

    print(f"写真重複チェック:{picture_data}")
    context = {'object': table_object, 'article_pk': article_pk, 'grouped_data': grouped_data, 'photo_grouped_data': photo_grouped_data, 'buttons': buttons, 'picture_data': picture_data}
    # 渡すデータ：　損傷データ　↑　　　       　   joinと損傷座標毎にグループ化したデータ　↑　　　　　　 写真毎にグループ化したデータ　↑ 　　       径間ボタン　↑
    # print(f"写真格納場所確認：{context}")
    # テンプレートをレンダリング
    return render(request, 'infra/bridge_table.html', context)
print("かかった時間_time1: ", time.time() - start1 )
# << entity_extension 　　　　関数をdxf_file.py(別モジュール)に移動 >>

# << find_square_around_text 関数をdxf_file.py(別モジュール)に移動 >>

# << create_picturelist 　　 関数をtasks.py    (非同期処理) に移動 >>


def handle_uploaded_file(f):
    import os
    from django.conf import settings
    from django.core.files.storage import FileSystemStorage
    
    fs = FileSystemStorage()
    filename = fs.save(f.name, f)
    print(f"filename：{filename}")
    folder_name = os.path.splitext(f.name)[0] # os.path.splittext：ファイルの拡張子を除いたベースネームを取得
    full_path = os.path.join(fs.location, filename)
    print(f"folder_name：{folder_name}")

    folder_name = os.path.dirname(full_path)
    print(f"folder_name：{folder_name}")
    return f'infra/img/{folder_name}/{filename}'
    # return os.path.join(settings.MEDIA_URL, filename)

# << 写真の変更内容を反映 >>
def upload_picture(request, article_pk, pk):
    bucket_name = "infraprotect"
    
    article = get_object_or_404(FullReportData, id=article_pk)
    print(article)
    infra = get_object_or_404(FullReportData, id=pk)
    print(infra)
    object_name = f"{article}/{infra}/"
    
    if request.method == 'POST':
        action = request.POST.get('action')
        bridge_id = request.POST.get('bridgeId')
        bridge = get_object_or_404(FullReportData, id=bridge_id)
        print("写真帳の変更を行います")
        print(f"action：{action}")
        print(f"bridge_id：{bridge_id}")
        print(f"bridge：{bridge}")
        
        if action == 'change':
            print(bridge.this_time_picture)
            
            old_picture_path = request.POST.get('oldPicturePath')
            new_picture_path = f"https://{bucket_name}.s3.amazonaws.com/{object_name}"
            # もし、↑と↓のprint文で変わらないのであれば、このreplaceに問題がある。
            
            print("変更する動作")
            print(f"old_picture_path：{old_picture_path}")
            print(f"new_picture_path：{new_picture_path}")
            print(f"bridge.this_time_picture：{bridge.this_time_picture}")
            
            if new_picture_path:
                bridge.this_time_picture = new_picture_path
                bridge.save()
                print(bridge.this_time_picture)
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'message': 'Failed to upload to S3'})


        elif action == 'add':
            new_picture_path = handle_uploaded_file(request.FILES['file'])
            if bridge.this_time_picture:
                bridge.this_time_picture += f', {new_picture_path}'
            else:
                bridge.this_time_picture = new_picture_path
            bridge.save()
            print("追加する動作")
            print(f"new_picture_path：{new_picture_path}")
            print(f"bridge.this_time_picture：{bridge.this_time_picture}")
            return JsonResponse({'success': True})

        elif action == 'delete':
            picture_path = request.POST.get('picturePath')
            pictures = bridge.this_time_picture.split(', ')
            pictures.remove(picture_path)
            bridge.this_time_picture = ', '.join(pictures) if pictures else None
            bridge.save()
            print("削除する動作")
            print(f"picture_path：{picture_path}")
            print(f"pictures：{pictures}")
            print(f"bridge.this_time_picture：{bridge.this_time_picture}")
            return JsonResponse({'success': True})

        return JsonResponse({'success': False})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})



# << 所見一覧の作成 >>
def observations_list(request, article_pk, pk):
    context = {}
    print("所見ID確認")
    infra = Infra.objects.filter(id=pk).first()
    print(f"Infra:{infra}") # 旗揚げチェック(4)
    article = infra.article
    print(f"article:{article}") # お試し(2)
    table = Table.objects.filter(infra=infra.id, article=article.id).first()
    # table = Table.objects.filter(id=pk).first()
    print(f"table_name:{table}")

    bucket_name = 'infraprotect'
    print(bucket_name)
    folder_name = article.案件名+"/"
    print(folder_name)
    pattern = f'*{infra.title}*/{infra.title}.dxf'
    print(pattern)
    
    # 該当するオブジェクトを取得
    matched_objects = match_s3_objects_with_prefix(bucket_name, folder_name, pattern)

    if matched_objects:
        print(f"該当オブジェクト：{matched_objects}")
    else:
        print("ファイルが見つかりません")

    # 結果を表示
    for obj_key in matched_objects:
        encode_dxf_filename = f"https://{bucket_name}.s3.ap-northeast-1.amazonaws.com/{obj_key}"
    
    dxf_filename = urllib.parse.quote(encode_dxf_filename, safe='/:') # スラッシュとコロン以外をエンコード

    print(f"dxfファイルのデコードURLは：{encode_dxf_filename}")
    print(f"dxfファイルの絶対URLは：{dxf_filename}")
    
    # URLから径間番号を取得
    if "search_title_text" in request.GET:
        search_title_text = request.GET["search_title_text"]
    else:
        search_title_text = "1径間"

    second_search_title_text = "損傷図"
    
    # sorted_items = create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text)
    """"""
    # 全パーツデータを取得

    infra_name = table.infra.title
    print(f"infra_name:{infra_name}")
    parts_data = PartsNumber.objects.filter(infra=pk)
    print(f"parts_data:{parts_data}")
    
    article = Article.objects.filter(id=article_pk).first()
    infra = Infra.objects.filter(id=pk).first()
    table = Table.objects.filter(infra=pk).first()    
    
    material_replace_map = {
        "鋼": "S",
        "コンクリート": "C",
        "アスファルト": "A",
        "ゴム": "R",
        "その他": "X",
    }
    
    number_change = {
        '①': '腐食',
        '②': '亀裂',
        '③': 'ゆるみ・脱落',
        '④': '破断',
        '⑤': '防食機能の劣化',
        '⑥': 'ひびわれ',
        '⑦': '剥離・鉄筋露出',
        '⑧': '漏水・遊離石灰',
        '⑨': '抜け落ち',
        '⑩': '補修・補強材の損傷',
        '⑪': '床版ひびわれ',
        '⑫': 'うき',
        '⑬': '遊間の異常',
        '⑭': '路面の凹凸',
        '⑮': '舗装の異常',
        '⑯': '支承部の機能障害',
        '⑰': 'その他',
        '⑱': '定着部の異常',
        '⑲': '変色・劣化',
        '⑳': '漏水・滞水',
        '㉑': '異常な音・振動',
        '㉒': '異常なたわみ',
        '㉓': '変形・欠損',
        '㉔': '土砂詰まり',
        '㉕': '沈下・移動・傾斜',
        '㉖': '洗掘',
    }

    # lank_order = ['a', 'b', 'c', 'd', 'e']  # ランクの順序をリストで定義
    # def get_lank_value(damage_name):
    #     """damage_nameのランク部分を取得する"""
    #     if "-" in damage_name:
    #         return damage_name.split('-')[-1]
    #     return None

    # FullReportDataの準備
    damage_comments = defaultdict(lambda: {'damage_lanks': [], 'this_time_pictures': []})

    for part in parts_data:
        part_full_name = f"{part.parts_name} {part.symbol}{part.number}"
        span_number = part.span_number + '径間'
        print(f"partデータ：{part}")

        # FullReportDataから該当するデータを取得
        report_data_list = FullReportData.objects.filter(
            parts_name=part_full_name, # FullReportDataのparts_nameオブジェクトがpart_full_name(主桁 Mg0101)と同じ、かつ
            span_number=span_number, # FullReportDataのspan_numberオブジェクトがspan_number(1径間)と同じ、かつ
            infra=part.infra, # FullReportDataのinfraオブジェクトがpart.infraと同じ場合
            article=part.article
        )
        
        

        # 条件に満たすデータが存在するかチェック
        if report_data_list.exists():
            for report_data in report_data_list:
                print(f"report_data:{report_data}")
                # print(f"picture:{report_data.this_time_picture}")

                damage_list_material = "" # 空のdamage_list_materialを用意
                for m in part.material.all(): # part.materialの全データを取得し「m」変数に入れる
                    damage_list_material += m.材料 + "," # 「m」の材料フィールドを指定してdamage_list_materialに入れる
                    
                elements = damage_list_material.split(',')
                replaced_elements = [material_replace_map.get(element, element) for element in elements] # それぞれの要素を置換辞書に基づいて変換します
                damage_list_materials = ','.join(replaced_elements) # カンマで結合
                
                damage_name = report_data.damage_name.split('-')[0] if '-' in report_data.damage_name else report_data.damage_name
                if damage_name == "NON":
                    damage_name = damage_name
                elif damage_name[0] != '⑰':
                    damage_name = number_change[damage_name[0]]
                else:
                    damage_name = damage_name[1:] # ⑦剥離・鉄筋露出から先頭の一文字を省く

                damage_lank = report_data.damage_name.split('-')[1] if '-' in report_data.damage_name else report_data.damage_name
                
                # DamageListに必要なフィールドを含むインスタンスを作成
                # << 損傷一覧(Excel)用データ登録 >>
                damage_list_entry = DamageList(
                    parts_name = part.parts_name, # 主桁
                    symbol = part.symbol, # Mg
                    number = part.number, # 0101
                    material = damage_list_materials[:-1], # 最後のコンマが不要なため[-1:]（S,C）
                    main_parts = "〇" if part.main_frame else "", # 主要部材のフラグ
                    damage_name = damage_name, # 剥離・鉄筋露出
                    damage_lank = damage_lank, # d
                    span_number = part.span_number,
                    infra = part.infra,
                    article = part.article
                )
                
                try:
                    # DamageListインスタンスを保存
                    damage_list_entry.save()
                    
                except IntegrityError:
                    # 重複データがある場合の処理
                    print("保存しませんでした。")
                    # 必要に応じてログを記録したり、他の処理を追加したりできます
                    # continue  # 次のループに進む
        else:
            material_text = ""
            for m in part.material.all():
                material_text += m.材料 + ","
                
            elements = material_text.split(',')
            replaced_elements = [material_replace_map.get(element, element) for element in elements] # それぞれの要素を置換辞書に基づいて変換します
            damage_list_materials = ','.join(replaced_elements) # カンマで結合
            
            # FullReportDataに該当がないパーツも新しいモデルに登録
            no_damage_parts_number = DamageList(
                parts_name = part.parts_name, # 主桁
                symbol = part.symbol, # Mg
                number = part.number, # 0101
                material = damage_list_materials[:-1], # 最後のコンマが不要なため[-1:]（S,C）
                main_parts = "〇" if part.main_frame else "", # 主要部材のフラグ
                damage_name = "NON", # 剥離・鉄筋露出
                damage_lank = "a", # d
                span_number = part.span_number,
                infra = part.infra,
                article = part.article
            )
            
            try:
                # DamageListインスタンスを保存
                no_damage_parts_number.save()
                
            except IntegrityError:
                # 重複データがある場合の処理
                print("保存しませんでした。")
                # 必要に応じてログを記録したり、他の処理を追加したりできます
                # continue  # 次のループに進む
                    
    """所見用のクラス登録"""
    damage_comments = defaultdict(lambda: {'damage_lanks': [], 'this_time_pictures': []})

    for part in parts_data:
        part_full_name = f"{part.parts_name} {part.symbol}{part.number}"
        span_number = part.span_number + '径間'

        report_data_list = FullReportData.objects.filter(
            parts_name=part_full_name,
            span_number=span_number,
            infra=part.infra,
            article=part.article
        )
        
        main_parts_list_left = ["主桁", "PC定着部"]
        main_parts_list_right = ["横桁", "橋台"]
        main_parts_list_zero = ["床版"]
        
        if report_data_list.exists():
            for report_data in report_data_list:

                parts_name = f"{part.parts_name} {part.number}"
                
                # ルールに則り、部材番号を作成
                
                if any(word in parts_name for word in main_parts_list_left): # main_parts_list_leftリストと一致した場合
                    left = parts_name.find(" ")
                    number2 = parts_name[left+1:]
                    number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                    result_parts_name = parts_name[:left] + " " + number_part[:2]
                elif any(word in parts_name for word in main_parts_list_right): # main_parts_list_rightリストと一致した場合
                    right = parts_name.find(" ")
                    number2 = parts_name[right+1:]
                    number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                    result_parts_name = parts_name[:right] + " " + number_part[2:] if len(number_part) < 5 else number_part[2:]
                elif any(word in parts_name for word in main_parts_list_zero): # main_parts_list_zeroリストと一致した場合
                    right = parts_name.find(" ")
                    result_parts_name = parts_name[:right] + " 00"
                else:
                    right = parts_name.find(" ")
                    result_parts_name = parts_name[:right] + "00"

                damage_name = report_data.damage_name.split('-')[0] if '-' in report_data.damage_name else report_data.damage_name
                if damage_name == "NON":
                    damage_name = damage_name
                elif damage_name[0] != '⑰':
                    damage_name = number_change[damage_name[0]]
                else:
                    damage_name = damage_name[1:] 
                damage_lank = report_data.damage_name.split('-')[1] if '-' in report_data.damage_name else report_data.damage_name
                    
                    
                # 部材名と損傷名の組み合わせでデータを作成
                damage_comments[(result_parts_name, damage_name)]['damage_lanks'].append(damage_lank)
                damage_comments[(result_parts_name, damage_name)]['this_time_pictures'].append(report_data.this_time_picture)
                
                damage_comment_material = ""
                for m in part.material.all():
                    damage_comment_material += m.材料 + ","
                elements = damage_comment_material.split(',')
                replaced_elements = [material_replace_map.get(element, element) for element in elements]
                damage_comment_materials = ','.join(replaced_elements)
                print(f"replaced_elements:{replaced_elements}")
                print(f"damage_comment_materials:{damage_comment_materials}")

                damage_comments[(result_parts_name, damage_name)]['material'] = damage_comment_materials[:-1]
                damage_comments[(result_parts_name, damage_name)]['main_parts'] = "〇" if part.main_frame else ""
                damage_comments[(result_parts_name, damage_name)]['span_number'] = part.span_number
                damage_comments[(result_parts_name, damage_name)]['infra'] = part.infra
                damage_comments[(result_parts_name, damage_name)]['article'] = part.article
                
        else:
            damage_name = "NON"
            damage_lank = "a"
            
            parts_name = f"{part.parts_name} {part.number}"
            
            # ルールに則り、部材番号を作成
            if any(word in parts_name for word in main_parts_list_left): # main_parts_list_leftリストと一致した場合
                left = parts_name.find(" ")
                number2 = parts_name[left+1:]
                number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                result_parts_name = parts_name[:left] + " " + number_part[:2]
            elif any(word in parts_name for word in main_parts_list_right): # main_parts_list_rightリストと一致した場合
                right = parts_name.find(" ")
                number2 = parts_name[right+1:]
                number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                result_parts_name = parts_name[:right] + " " + number_part[2:] if len(number_part) < 5 else number_part[2:]
            elif any(word in parts_name for word in main_parts_list_zero): # main_parts_list_zeroリストと一致した場合
                right = parts_name.find(" ")
                result_parts_name = parts_name[:right] + " 00"
            else:
                right = parts_name.find(" ")
                result_parts_name = parts_name[:right] + "00"

                
            # 部材名と損傷名の組み合わせでデータを作成
            damage_comments[(result_parts_name, damage_name)]['damage_lanks'].append(damage_lank)
            damage_comments[(result_parts_name, damage_name)]['this_time_pictures'].append(report_data.this_time_picture)
            
            
            
            damage_comment_material = ""
            for m in part.material.all():
                damage_comment_material += m.材料 + ","
            elements = damage_comment_material.split(',')
            replaced_elements = [material_replace_map.get(element, element) for element in elements]
            damage_comment_materials = ','.join(replaced_elements)
            print(f"replaced_elements:{replaced_elements}")
            print(f"damage_comment_materials:{damage_comment_materials}")

            damage_comments[(result_parts_name, damage_name)]['material'] = damage_comment_materials[:-1]
            damage_comments[(result_parts_name, damage_name)]['main_parts'] = "〇" if part.main_frame else ""
            damage_comments[(result_parts_name, damage_name)]['span_number'] = part.span_number
            damage_comments[(result_parts_name, damage_name)]['infra'] = part.infra
            damage_comments[(result_parts_name, damage_name)]['article'] = part.article
            
    for (result_parts_name, damage_name), data in damage_comments.items():
        print("1571行目")
        damage_lanks = data['damage_lanks']
        damage_max_lank = max(damage_lanks)
        damage_min_lank = min(damage_lanks)
        
        start_comma_pictures = ','.join(str(picture) for picture in set(data['this_time_pictures']) if picture is not None) # 重複なし
        # double_comma_picturesにデフォルト値を設定
        double_comma_pictures = start_comma_pictures
        
        if start_comma_pictures.startswith(","):
            double_comma_pictures = start_comma_pictures[1:]
        before_combined_pictures = double_comma_pictures.replace(",,", ",")
        # print(f"before_combined_pictures：{before_combined_pictures}")

        # << 管理サイトに登録するコード（所見） >>
        print(f"result_parts_name：{result_parts_name}")
        print(f"damage_name：{damage_name}")
        print(f"damage_max_lank：{damage_max_lank}")
        print(f"damage_min_lank：{damage_min_lank}")
        try:
            damage_comment_entry, created = DamageComment.objects.get_or_create(
                parts_name=result_parts_name,
                damage_name=damage_name,
                span_number=data['span_number'],
                infra=data['infra'],
                article=data['article'],
                defaults={
                    'material': data['material'],
                    'main_parts': data['main_parts'],
                    'damage_max_lank': damage_max_lank,
                    'damage_min_lank': damage_min_lank,
                    'this_time_picture': before_combined_pictures
                }
            )
            if not created:
                # 既存データが見つかった場合には、フィールド値を更新
                damage_comment_entry.material = data['material']# 材料
                damage_comment_entry.main_parts = data['main_parts']# 主要部材
                damage_comment_entry.damage_max_lank = damage_max_lank
                damage_comment_entry.damage_min_lank = damage_min_lank
                damage_comment_entry.this_time_picture = before_combined_pictures
                damage_comment_entry.save()

        except IntegrityError:
            # 重複データがある場合の処理
            print("データが存在しています。")
            # 必要に応じてログを記録したり、他の処理を追加したりできます
            continue  # 次のループに進む
        
        # BridgePictureからのimageを取得し、damage_comment_entryに追加する
        print(f"BridgePictureチェック：{result_parts_name}")
        print(f"BridgePictureチェック：{damage_name}")
        print(report_data.parts_split)
        print(report_data.damage_name)
        
        bridge_pictures = BridgePicture.objects.filter(
            memo__contains=f"{result_parts_name},{damage_name}", # __contains：部分一致
            span_number=f"{data['span_number']}径間",
            infra=data['infra'],
            article=data['article']
        )

        # def remove_unwanted_prefix(text):
        #     # 「https:/」を探して、それ以前の部分を削除する
        #     parts = text.split('https:/', 1)
        #     if len(parts) > 1:
        #         return 'https:/' + parts[1]
        #     else:
        #         return None
        
        images = []
        for picture in bridge_pictures:
            decoded_url = unquote(picture.image.url)
            print(f"写真パス：{decoded_url}")
            images.append(decoded_url)
            # images.append(remove_unwanted_prefix(decoded_url))
            print(picture)

        damage_comment_entry.images = images
    
    if "search_title_text" in request.GET:
        search_title_text = request.GET["search_title_text"]

    else:
        search_title_text = "1径間"
    
    span_create_number = search_title_text.replace("径間", "")
    print(span_create_number)
    filtered_bridges = DamageComment.objects.filter(infra=pk, span_number=span_create_number).order_by('span_number', 'replace_name', 'parts_number', 'number')
    # まず、重複しているデータを特定します。
    duplicates = DamageComment.objects.values(
        'infra', 'span_number', 'replace_name', 'parts_number'
    ).annotate(
        count=Count('id')
    ).filter(
        count__gt=1
    )

    # 各重複データについて、numberが27のものを削除します。
    for duplicate in duplicates:
        infra = duplicate['infra']
        span_number = duplicate['span_number']
        replace_name = duplicate['replace_name']
        parts_number = duplicate['parts_number']

        # 重複データのうち、numberが27のものをフィルタリング
        target_comments = DamageComment.objects.filter(
            infra=infra,
            span_number=span_number,
            replace_name=replace_name,
            parts_number=parts_number,
            number=27
        )

        # 存在する場合は削除
        if target_comments.exists():
            target_comments.delete()
            print(f"Deleted duplicate comments with number 27 for infra={infra}, span_number={span_number}.")

    # 並び替え時に「部材が重複」かつ「numberが27(NON)」の場合、そのデータを削除
    print(f"bridges:{filtered_bridges}")
    buttons_count = int(table.infra.径間数) # 数値として扱う
    buttons = list(range(1, buttons_count + 1)) # For loopのためのリストを作成
    # range(一連の整数を作成):range(1からスタート, ストップ引数3 = 2 + 1) → [1, 2](ストップ引数は含まれない)
    print(buttons)

    # print(f"所見ボタン:{DamageComment.objects.filter(infra=pk)}")# ボタン:<QuerySet [<Table: Table object (15)>]>
    # print(f"所見ボタン:{DamageComment.objects.filter(infra=pk).first()}")# ボタン:Table object (18)(QuerySetのままだとうまく動作しない)
    #   1(径間)  ,      1(主桁)  ,        01     ,    6(ひびわれ)

    # print("所見引き渡しID確認")
    infra = Infra.objects.filter(id=pk).first()
    article = infra.article
    observer_object = infra
    
    images_url = [] # ここで毎回初期化されます
    for observer_data in filtered_bridges:
        print(f"損傷名：{observer_data.damage_name}")
        
        if observer_data.damage_name.startswith("その他"):
            damage_name_result = "その他"
        else:
            pattern = r'^[①-㉖]|-[a-zA-Z]$'# 「その他」の場合、「の他」と変換されてしまう
            damage_name_result = re.sub(pattern, '', observer_data.damage_name)
        print(f"123:{damage_name_result}")
        
        for key, value in number_change.items():
            if value == damage_name_result:# 「その他(分類)」では番号は見つからない。そのため ↑ で「その他」としている
                print(f"123:{damage_name_result}")
                # 一致する場合、丸番号を text の先頭に追加
                number_and_text = key + damage_name_result
        print(f"number_and_text:{number_and_text}")
        
        span_number_data = observer_data.span_number + "径間"
        # クエリセットでフィルタリング
        print(f"observer_data.damage_name:{observer_data.damage_name}")
        print(observer_data.parts_name)
        print(damage_name_result)
        print(f"{observer_data.parts_name},{number_and_text}")
        print("")
        if " " not in observer_data.parts_name:
            observer_data.parts_name = re.sub(r'(\D)(\d)', r'\1 \2', observer_data.parts_name)
            print(observer_data.parts_name)
            
        matches = BridgePicture.objects.filter( # 部材名/部材番号/損傷名/径間名/案件名/橋梁名
            parts_split=observer_data.parts_name, # 主桁 01/主桁 01
            damage_name__icontains=damage_name_result, # ⑦剥離・鉄筋露出-e/剥離・鉄筋露出
            #memo__icontains=f"{observer_data.parts_name},{number_and_text}",# 排水管 00,①腐食(小大)-c,
            span_number=span_number_data, # 1径間/1
            infra=observer_data.infra,
            article=observer_data.article
        ).distinct()
        if not matches:
            matches = BridgePicture.objects.filter( # 部材名/部材番号/損傷名/径間名/案件名/橋梁名
                #parts_split=observer_data.parts_name, # 主桁 01/主桁 01
                #damage_name__icontains=damage_name_result, # ⑦剥離・鉄筋露出-e/剥離・鉄筋露出
                memo__icontains=f"{observer_data.parts_name},{number_and_text}",# 排水管 00,①腐食(小大)-c,
                span_number=span_number_data, # 1径間/1
                infra=observer_data.infra,
                article=observer_data.article
            )
        
        match_details = [
            {"id": match.picture_number, "other_field": match.image}
            for match in matches
        ]
            
        images_url.append({"full_report": observer_data, "matches": matches, "match_details": match_details})
        
    context = {'object': observer_object, 'article_pk': article_pk, 'data': filtered_bridges, 'article_pk': article_pk, 'pk': pk, 'buttons': buttons, 'images': images_url}
    # print(f"所見用context：{context}")
    return render(request, 'infra/observer_list.html', context)


# << 所見コメントのリアルタイム保存 >>
def damage_comment_edit(request, pk):
    if request.method == "POST":
        # TODO: 編集を受け付ける
        # DamageComment の idを受け取る。
        # URL：path('damage_comment_edit/<int:pk>/', views.damage_comment_edit , name="damage_comment_edit")
        damage_comment = DamageComment.objects.get(id=pk) # idが同じDamageCommentデータを取得(int:pk 1種類のidが必要)
        print(damage_comment)
        form = DamageCommentEditForm(request.POST, instance=damage_comment)
     # ユーザーが送信したPOSTデータをFormに渡す ↑　　　　　　　　↑ 編集するオブジェクト
        print("編集します。")

        if form.is_valid(): # バリデーション
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)
        # リダイレクト処理  　　　　　　　　　　　　　↓　damage_commentクラス → infraフィールド(infraクラスに移る) → articleフィールド(articleクラスに移る)からarticle.idを取得
        return redirect("observations-list", damage_comment.infra.article.id, damage_comment.infra.id )
    
    
# << どの対策区分ボタンが押されたか、管理サイトに保存 >>
def damage_comment_jadgement_edit(request, pk):
    if request.method == "POST":
        #TODO: 編集を受け付ける。
        # DamageComment の idを受け取る。
        print("DamageCommentJadgementEditForm 発動。")
        damage_comment = DamageComment.objects.get(id=pk)
        form = DamageCommentJadgementEditForm(request.POST, instance=damage_comment)
        
        if form.is_valid():
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)

        return redirect("observations-list", damage_comment.infra.article.id, damage_comment.infra.id )
    
# << どの損傷原因ボタンが押されたか、管理サイトに保存 >>
def damage_comment_cause_edit(request, pk):
    if request.method == "POST":
        #TODO: 編集を受け付ける。
        # DamageComment の idを受け取る。
        print("DamageCommentCauseEditForm 発動。")
        damage_comment_cause = DamageComment.objects.get(id=pk)
        form = DamageCommentCauseEditForm(request.POST, instance=damage_comment_cause)
        
        if form.is_valid():
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)

        return redirect("observations-list", damage_comment_cause.infra.article.id, damage_comment_cause.infra.id )
    
    
# << エクセル出力時に並び替えを行う >>
parts_name_priority_list = ['主桁', '横桁', '床版']
# カスタムソートキー関数（エクセル出力時に使用）
def custom_sort_key(record):
    # parts_nameリストの優先度を求めるためのインデックス
    parts_name_priority = next((i for i, part in enumerate(parts_name_priority_list) if part in record.parts_name), len(parts_name_priority_list))
    return (int(record.span_number), parts_name_priority)

# << 管理サイトに登録したデータをエクセルに出力 >>
def excel_output(request, article_pk, pk):
    used_image_list = []
    bridge_name = ""

    bucket_name = 'infraprotect'
    original_file_path = f"https://{bucket_name}.s3.ap-northeast-1.amazonaws.com/H31_bridge_base.xlsm"
    print(f"エクセルひな形のURL：{original_file_path}")
    # エクセルファイルを読み込む
    s3 = boto3.client('s3')
    response = s3.get_object(Bucket=bucket_name, Key="H31_bridge_base.xlsm")
    file_stream = BytesIO(response['Body'].read())
    wb = openpyxl.load_workbook(file_stream, keep_vba=True)
    print("読み取り")
    # << Django管理サイトからデータを取得（その１用） >>
    no01_records = Infra.objects.filter(id=pk, article=article_pk)
    ws = wb['その１']
    for record in no01_records:
        bridge_name = record.title
        print(bridge_name)
        ws[f'F6'] = record.title # 〇〇橋
        ws[f'O10'] = record.橋長
        ws[f'BC5'] = record.橋梁コード
        ws[f'BF11'] = record.交通量
        ws[f'BF13'] = record.大型車混入率
        # 活荷重を処理
        load_weights = ', '.join([str(weight) for weight in record.活荷重.all()])
        ws['X10'] = load_weights
        # 等級を処理
        load_grades = ', '.join([str(grade) for grade in record.等級.all()])
        ws['AD10'] = load_grades
        # 適用示方書を処理
        rulebooks = ', '.join([str(rulebook) for rulebook in record.適用示方書.all()])
        ws['AK10'] = rulebooks

    # << Django管理サイトからデータを取得（その７、８用） >>
    no0708_records = DamageComment.objects.filter(infra=pk, article=article_pk)
    # 並び替えて出力
    sorted_records = sorted(no0708_records, key=custom_sort_key)
    # カウンタ変数をシートごとに用意
    span = 1
    i07, i08 = 0, 0
    initial_row07, initial_row08= 13, 13 # 1つ目の入力位置
    
    for record in sorted_records:
        print(f"出力レコード:{record}")
        print(f"　径間:{span}")
        if int(record.span_number) == span + 1:
            span = int(record.span_number)
            initial_row07 = initial_row07 + 8 * math.ceil(i07 / 8) # 13+8×(ページ数)
            initial_row08 = initial_row08 + 8 * math.ceil(i08 / 8) # 13,21,29(+8)
            i07, i08 = 0, 0
        if int(record.span_number) == span:
            if record.main_parts == "〇":
                ws = wb['その７']
                row = initial_row07 + i07 # 行は13から
                i07 += 1
            else:
                ws = wb['その８']
                row = initial_row08 + i08 # 行は13から
                i08 += 1
            print(f"　エクセル出力:{record.comment_parts_name}{record.parts_number}{record.damage_name}{record.jadgement}")
            ws[f'F{row}'] = record.comment_parts_name # 主桁
            ws[f'J{row}'] = record.parts_number # 01
            ws[f'D{row}'] = record.material # S,C
            ws[f'L{row}'] = record.damage_max_lank # e
            ws[f'N{row}'] = record.damage_min_lank # b
            ws[f'BO{row}'] = record.span_number # 径間番号
            
            if record.damage_name != "NON":
                if record.jadgement == "C1":
                    jadgement_position = f'S{row}'
                elif record.jadgement == "C2":
                    jadgement_position = f'V{row}'
                elif record.jadgement == "M":
                    jadgement_position = f'Z{row}'
                elif record.jadgement == "E1":
                    jadgement_position = f'AD{row}'
                elif record.jadgement == "E2":
                    jadgement_position = f'AH{row}'
                elif record.jadgement == "S1":
                    jadgement_position = f'AK{row}'
                elif record.jadgement == "S2":
                    jadgement_position = f'AN{row}'
                else:                  # "B"
                    jadgement_position = f'P{row}'
            
                ws[jadgement_position] = record.damage_name # 腐食
                ws[f'AU{row}'] = record.cause # 損傷原因「経年変化」  
                print(f"初見コメント：{record.comment}")
            
                if record.comment != None:
                    choice_comment = record.comment
                else:
                    choice_comment = record.auto_comment
                
                ws[f'BC{row}'] = choice_comment # 〇〇が見られる。
            else:
                ws[f'BC{row}'] = "健全である。"

    # << （その１０） >>
    no10_records = FullReportData.objects.filter(infra=pk, article=article_pk, this_time_picture__isnull=False).exclude(this_time_picture='')
    #                                                                          this_time_fieldがnull(空)=でない 除外する(this_time_picture='')
    ws = wb['その１０']
    print(f"最大径間数：{span}")
    print(f"最大径間数：{int(span)}")
    # << セル位置を作成 >>
    picture_and_spannumber_row = 9 # 部材名・要素番号
    partsname_and_number_row = 10 # 部材名・要素番号
    damagename_and_lank_row = 11 # 損傷の種類・損傷程度
    picture_start_row = 13 # 損傷写真
    lasttime_lank_row = 15 # 前回損傷程度
    damage_memo_row = 17 # 損傷メモ
    step = 14
    output_data = len(no10_records)
    num_positions = math.ceil(output_data/3) + int(span) * 6 # 横3列で割って何行になるか
    
    # 関連する列を定義
    picture_columns = ["E", "AE", "BE"] # 写真列
    left_columns = ["I", "AI", "BI"] # 左列
    right_columns = ["R", "AR", "BR"] # 右列
    bottom_columns = ["T", "AT", "BT"] # 前回程度+メモ
    # セル位置のリストを生成
    join_picturenumber_cell = [f"{col}{picture_and_spannumber_row + i * step}" for i in range(num_positions) for col in left_columns] # 写真番号
    join_spannumber_cell = [f"{col}{picture_and_spannumber_row + i * step}" for i in range(num_positions) for col in right_columns] # 径間番号
    join_partsname_cell = [f"{col}{partsname_and_number_row + i * step}"    for i in range(num_positions) for col in left_columns] # 部材名
    join_number_cell = [f"{col}{partsname_and_number_row + i * step}"       for i in range(num_positions) for col in right_columns] # 要素番号
    join_damagename_cell = [f"{col}{damagename_and_lank_row + i * step}"    for i in range(num_positions) for col in left_columns] # 損傷の種類
    join_lank_cell = [f"{col}{damagename_and_lank_row + i * step}"          for i in range(num_positions) for col in right_columns] # 損損傷程度
    join_picture_cell = [f"{col}{picture_start_row + i * step}"             for i in range(num_positions) for col in picture_columns] # 損傷写真
   #join_lasttime_lank_cell = [f"{col}{lasttime_lank_row + i * step}"       for i in range(num_positions) for col in bottom_columns] # 前回損傷程度
    join_damage_memo_cell = [f"{col}{damage_memo_row + i * step}"           for i in range(num_positions) for col in bottom_columns] # 損傷メモ

    span = 1
    page_count = 1
    i10 = 0
    initial_row10 = 9 # 1つ目の入力位置
    
    def hide_sheet_copy_and_paste(wb, sheet_name):
        """シートを再表示してコピーその後非表示に設定"""

        hide_sheet = wb['ページ１０']
        hide_sheet.sheet_state = 'visible'

        # コピーする行の範囲を指定します
        copy_start_row = 2
        copy_end_row = 29

        # コピーする行のデータとスタイルを保持するリストを作成します
        rows_to_copy = []
        merges_to_keep = []

        for row_idx in range(copy_start_row, copy_end_row + 1):
            row_data = []
            for cell in hide_sheet[row_idx]:
                cell_data = {
                    'value': cell.value,
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection),
                    'alignment': copy(cell.alignment)
                }
                row_data.append(cell_data)
            row_data.append(hide_sheet.row_dimensions[row_idx].height)
            rows_to_copy.append(row_data)

        # 元のシートのセル結合情報を取得
        for merge in hide_sheet.merged_cells.ranges:
            if (copy_start_row <= merge.min_row <= copy_end_row) or \
                (copy_start_row <= merge.max_row <= copy_end_row):
                merges_to_keep.append(copy(merge))
        
        sheet = ws
        
        # コピー先の行を挿入します
        # A列の一番下の行番号を取得
        max_row = sheet.max_row
        while sheet['A' + str(max_row)].value is None and max_row > 0:
            max_row -= 1
        insert_at_row = max_row
        # print(f"max_row：{max_row}")
        
        # シフトする行の高さを保持するリストを作成します
        heights = []
        for row_idx in range(insert_at_row, sheet.max_row + 1):
            heights.append(sheet.row_dimensions[row_idx].height)
        
        # 指定行から下の行をシフト
        sheet.insert_rows(insert_at_row, amount=(copy_end_row - copy_start_row + 1))

        # 行の高さを元に戻す
        for i, height in enumerate(heights):
            sheet.row_dimensions[insert_at_row + i + (copy_end_row - copy_start_row + 1)].height = height

        # コピーされた行を挿入
        for i, row_data in enumerate(rows_to_copy):
            new_row = insert_at_row + i
            for j, cell_data in enumerate(row_data[:-1]):
                cell = sheet.cell(row=new_row, column=j + 1)
                cell.value = cell_data['value']
                cell.font = cell_data['font']
                cell.border = cell_data['border']
                cell.fill = cell_data['fill']
                cell.number_format = cell_data['number_format']
                cell.protection = cell_data['protection']
                cell.alignment = cell_data['alignment']
            sheet.row_dimensions[new_row].height = row_data[-1]

        # セル結合をコピー
        for merged_range in merges_to_keep:
            new_min_row = merged_range.min_row - copy_start_row + insert_at_row
            new_max_row = merged_range.max_row - copy_start_row + insert_at_row
            new_merge_range = "{}{}:{}{}".format(
                openpyxl.utils.get_column_letter(merged_range.min_col),
                new_min_row,
                openpyxl.utils.get_column_letter(merged_range.max_col),
                new_max_row
            )
            sheet.merge_cells(new_merge_range)
            
        # 最大行を取得
        max_row = sheet.max_row
        # 印刷範囲の設定を修正
        start_col = "A"
        end_col = 'CD'
        print_area = f"{start_col}1:{end_col}{max_row}"
        sheet.print_area = print_area    
        
        hide_sheet.sheet_state = 'hidden'   
                    
    # 全ての結果を入れるリスト
    joined_results = []

    # FullReportDataをリスト化
    no10_records_list = list(no10_records)
    # 座標によってFullReportDataを辞書に格納
    full_report_dict = defaultdict(list)
    for record in no10_records_list:
        key = (record.damage_coordinate_x, record.damage_coordinate_y, record.table, record.infra, record.article)
        full_report_dict[key].append(record)
        
    damage_picture_data = BridgePicture.objects.filter(infra=pk, article=article_pk)
    # BridgePictureのデータを処理
    for picture in damage_picture_data:
        key = (picture.damage_coordinate_x, picture.damage_coordinate_y, picture.table, picture.infra, picture.article)
        matching_records = full_report_dict.get(key) # キー一致のレコードを取得
        print(picture.image)
        
        if matching_records:
            print("チェック")
            for record in matching_records:
                combined_result = {
                    'parts_name': record.parts_name,
                    'damage_name': record.damage_name,
                    'span_number': record.span_number,
                    'textarea_content': record.textarea_content,
                    'image': picture.image,
                    'picture_number': picture.picture_number
                }
                # imageが空でない場合にのみ追加
                if combined_result['image']:
                    joined_results.append(combined_result) # 管理サイトにデータが格納されているか確認
    # print(f"ここは合体したデータ{len(joined_results)}個：{joined_results}")            

    i10 = 0
    unique_combinations = set()
    for record in joined_results:
        span_number = record['span_number'].replace('径間', '')
        print(f"ここは{span_number}径間目")
        print(f"その10レコード数：{len(no10_records)}")
        if int(span_number) == span + 1:
            hide_sheet_copy_and_paste(wb, ws)
            span = int(span_number)
            print(f"－－－{span}径間に変更")            

            page_plus = math.ceil(i10/6)
            print(f"現在、{page_plus}ページ目")
            i10 = page_plus * 6
            print(f"径間が変わるとしたら{i10}個目")
            ws[join_spannumber_cell[i10]] = span
            
        if int(span_number) == span: # 同じ径間の場合
            if i10 % 6 == 5 and i10 > 10:
                hide_sheet_copy_and_paste(wb, ws) # プログラム4｜1ページ増やすマクロを実行

            print(f"対象数:{i10}/{len(no10_records)}")
            # 部材名を入力形式に分ける( 主桁 0101 )
            if " " in record['parts_name']:
                split_parts = record['parts_name'].split(" ")
                parts_name = split_parts[0]
                print(f"1-parts_name：{parts_name}") # 防護柵
                parts_number = re.search(r'\d+', split_parts[1]).group()
                print(f"2-parts_number：{parts_number}") # 0101
            else:
                print("カッコなし")
                parts_name = ""
                parts_number = ""
            # 損傷名を入力形式に分ける( ⑦剥離・鉄筋露出-e )
            if "-" in record['damage_name']:
                # 置き換え用の辞書
                number_change = {
                    '①': '腐食',
                    '②': '亀裂',
                    '③': 'ゆるみ・脱落',
                    '④': '破断',
                    '⑤': '防食機能の劣化',
                    '⑥': 'ひびわれ',
                    '⑦': '剥離・鉄筋露出',
                    '⑧': '漏水・遊離石灰',
                    '⑨': '抜け落ち',
                    '⑩': '補修・補強材の損傷',
                    '⑪': '床版ひびわれ',
                    '⑫': 'うき',
                    '⑬': '遊間の異常',
                    '⑭': '路面の凹凸',
                    '⑮': '舗装の異常',
                    '⑯': '支承部の機能障害',
                    '⑰': 'その他',
                    '⑱': '定着部の異常',
                    '⑲': '変色・劣化',
                    '⑳': '漏水・滞水',
                    '㉑': '異常な音・振動',
                    '㉒': '異常なたわみ',
                    '㉓': '変形・欠損',
                    '㉔': '土砂詰まり',
                    '㉕': '沈下・移動・傾斜',
                    '㉖': '洗掘',
                }

                first_char = record['damage_name'][0] # 先頭の1文字を取得                
                print(f"first_char　{first_char}")
                damage_name = number_change.get(first_char, "") # 辞書で値を取得
                damage_lank = record['damage_name'][-1]
                print(f"3-damage_name　{damage_name}") # 損傷種類（ ひびわれ ）
                print(f"4-damage_lank　{damage_lank}") # 損傷程度（    d    ）
                print(f"5-picture_number　{record['picture_number']}")
            else:
                damage_name = ""
                print(f"3-damage_nameなし") # 損傷種類（ ひびわれ ）
                damage_lank = ""
                print(f"4-damage_lankなし") # 損傷程度（    d    ）     
                          
            print(f"damage_picture_data：{record['image']}")
            # 最大の写真サイズ（幅、高さ）
            
            combination = (record['picture_number'], record['image'], record['span_number'])
            if combination in unique_combinations:
                print(f"Duplicate entry found: {combination}")
                continue  # Skip duplicate entry
            else:
                unique_combinations.add(combination)
                
            max_width, max_height = 240, 180 # 4:3
            ws[join_picturenumber_cell[i10]] = record['picture_number'] # 写真番号
            ws[join_partsname_cell[i10]] = parts_name # 部材名
            ws[join_number_cell[i10]] = parts_number # 要素番号
            ws[join_damagename_cell[i10]] = damage_name # 損傷の種類
            ws[join_lank_cell[i10]] = damage_lank # 損傷の程度
                # 写真の有無を判断
            try:
                image_path = record['image'] # ImageFieldの場合は.pathをつける
                print(image_path)
                used_image_list.append(image_path)# 見つけた写真URLをリストに格納
                
            except AttributeError:
                print(f"エントリに 'this_time_picture' が存在しないか、無効です")
                continue  # このエントリをスキップ
            
            # image_path = os.path.join(settings.BASE_DIR, record['image'].lstrip('/')) # フルパスに変更
            # print(f"Calculated image path: {image_path}") 
            
            # TODO：openpyxlはローカル写真の貼付けのみ、S3バケットの写真をDLせず貼付けることは不可能
            # 画像をS3からダウンロードしてメモリ上に保存
            try:
                response = requests.get(image_path) # 指定したURLから画像データを取得
                response.raise_for_status()  # リクエストが成功したか確認(失敗したら例外処理となる)

                image_data = response.content
                pil_img = PILImage.open(io.BytesIO(image_data))
                
                width, height = pil_img.size
                aspect_ratio = width / height

                if aspect_ratio > max_width / max_height:
                    new_width = min(width, max_width)
                    new_height = new_width / aspect_ratio
                else:
                    new_height = min(height, max_height)
                    new_width = new_height * aspect_ratio

                resized_img = pil_img.resize((int(new_width), int(new_height)))
                
                # 画像を一時ファイルとして保存
                with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp:
                    resized_img_path = tmp.name# 写真のリサイズ
                    resized_img.save(resized_img_path)# 一時ファイルに保存
                    img = OpenpyxlImage(resized_img_path)

                img.anchor = ws[join_picture_cell[i10]].coordinate
                ws.add_image(img)
                
            except requests.exceptions.RequestException as e:
                print(f"写真貼付けエラー: {e}")
                
            ws[join_damage_memo_cell[i10]] = record['textarea_content'] # メモ
            print(record['textarea_content'])
            i10 += 1
            print(i10)

    # << Django管理サイトからデータを取得（その１１、１２用） >>
    no1112_records = DamageList.objects.filter(infra=pk, article=article_pk)
    # 並び替え
    sorted_records = sorted(no1112_records, key=custom_sort_key)
    span = 1
    i11, i12 = 0, 0
    initial_row11, initial_row12 = 10, 10 # 1つ目の入力位置

    for record in sorted_records:
        print(f"出力レコード:{record}")
        print(f"　径間:{span}")
        if int(record.span_number) == span + 1:
            span = int(record.span_number)
            initial_row11 = initial_row11 + 18 * math.ceil(i11 / 18) # 10+18×(ページ数)
            initial_row12 = initial_row12 + 18 * math.ceil(i12 / 18) # 10,28,46(+18)
            i11, i12 = 0, 0
        if int(record.span_number) == span:
            if record.main_parts == "〇":
                ws = wb['その１１']
                row = initial_row11 + i11 # 行は10から
                i11 += 1
            else:
                ws = wb['その１２']
                row = initial_row12 + i12 # 行は10から
                i12 += 1
            print(f"　エクセル出力:{record.parts_name}{record.damage_name}{record.span_number}")
            ws[f'H{row}'] = record.parts_name # 主桁
            ws[f'T{row}'] = record.number # 0101
            ws[f'E{row}'] = record.material # S,C
            ws[f'AR{row}'] = record.damage_name # 腐食
            ws[f'X{row}'] = record.damage_lank # d
            ws[f'BE{row}'] = record.classification # 分類「1」
            ws[f'AO{row}'] = record.pattern # パターン「6」
            ws[f'BL{row}'] = record.span_number # 径間番号

    # 現在の日時を取得してファイル名に追加
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    # 新しいファイル名の生成
    new_filename = f"{bridge_name}(作成：{timestamp}).xlsm"# _{original_file_path}"
    # サンプル橋(作成：20241015_114539)_base.xlsm
    
    # 画像をダウンロードしてZIP化
    zip_data = download_and_zip_images(used_image_list)

    # ダウンロードフォルダのパスを取得
    download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    # ファイルパスを作成
    file_path = os.path.join(download_folder, f'使用写真(DL：{timestamp}).zip')

    # ZIPデータをPCのダウンロードフォルダに保存
    with open(file_path, 'wb') as f:
        f.write(zip_data.read())

    print("Images have been zipped into 'downloaded_images.zip'")
    
    #メモリ空間内に保存
    virtual = BytesIO()
    wb.save(virtual)
    #バイト文字列からバイナリを作る
    binary = BytesIO(virtual.getvalue())
    return FileResponse(binary, filename = new_filename)